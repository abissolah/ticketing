from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView as BaseLoginView
from django.core.paginator import Paginator
from django.db.models import Case, Count, IntegerField, Q, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.views.generic import CreateView, DetailView, UpdateView, ListView
from django_filters.views import FilterView
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .filters import TicketFilter
from .forms import TicketCreateForm, TicketForm, TicketCommentForm
from .models import (
    Ticket,
    TicketComment,
    TicketAttachment,
    ClientMember,
    STATUS_CHOICES,
)
from .utils import (
    get_user_client,
    get_user_collaborateur,
    get_visible_tickets_queryset,
    get_all_tickets_queryset,
    get_archived_tickets_queryset,
    can_create_ticket,
    can_manage_clients_collaborateurs,
    get_clients_for_collaborateur,
    get_ticket_unread_comment_count,
    get_tickets_unread_counts,
)


class LoginView(BaseLoginView):
    template_name = "tickets/login.html"
    redirect_authenticated_user = True


def logout_view(request):
    auth_logout(request)
    return redirect("tickets:login")


class TicketListView(LoginRequiredMixin, FilterView):
    """Liste des tickets (tableau ou tuiles) avec filtres et export Excel."""
    model = Ticket
    filterset_class = TicketFilter
    template_name = "tickets/ticket_list.html"
    paginate_by = 25
    context_object_name = "tickets"

    def get_queryset(self):
        qs = get_all_tickets_queryset(self.request.user)
        # Par défaut : uniquement les tickets non archivés ; inclure les archivés si filtre "Archivés" est choisi
        archived_param = self.request.GET.get("archived", "").lower()
        if archived_param not in ("true", "1"):
            qs = qs.filter(archived=False)
        qs = qs.select_related(
            "client", "member", "assigned_to", "assigned_to__user"
        )
        return _order_tickets_priority_then_oldest(qs)

    def get_filterset_kwargs(self, filterset_class):
        kwargs = super().get_filterset_kwargs(filterset_class)
        client = get_user_client(self.request.user)
        collab = get_user_collaborateur(self.request.user)
        if client:
            kwargs["client_ids"] = [client.id]
        elif collab:
            ids_prestataire = set(
                collab.prestataire.clients.values_list("id", flat=True)
            )
            ids_collab = set(collab.clients.values_list("id", flat=True))
            kwargs["client_ids"] = list(ids_prestataire | ids_collab)
            kwargs["collaborateur"] = collab
        else:
            kwargs["client_ids"] = []
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["view_mode"] = self.request.GET.get("view", "table")
        ctx["priority_colors"] = {
            "low": "#198754",
            "medium": "#fd7e14",
            "high": "#dc3545",
        }
        ctx["status_choices"] = STATUS_CHOICES
        ctx["can_create"] = can_create_ticket(self.request.user)
        q = self.request.GET.copy()
        q.pop("page", None)
        ctx["query_string"] = q.urlencode()
        q_pop_view = self.request.GET.copy()
        q_pop_view.pop("page", None)
        q_pop_view.pop("view", None)
        ctx["query_string_no_view"] = q_pop_view.urlencode()
        # Pour qu'un collaborateur puisse affecter un ticket depuis la liste
        collab = get_user_collaborateur(self.request.user)
        ctx["assignable_collaborateurs"] = (
            collab.prestataire.collaborateurs.order_by("last_name", "first_name")
            if collab else None
        )
        # Tickets archivés (autre tableau, limité à 100)
        archived_qs = get_archived_tickets_queryset(self.request.user)
        archived_qs = archived_qs.select_related(
            "client", "member", "assigned_to", "assigned_to__user"
        )
        ctx["archived_tickets"] = _order_tickets_priority_then_oldest(archived_qs)[:100]
        # Nombre de commentaires non lus par ticket (pour la pastille)
        tickets_on_page = ctx.get("tickets") or ctx.get("object_list") or []
        ticket_ids = [t.id for t in tickets_on_page]
        ctx["ticket_unread_counts"] = get_tickets_unread_counts(ticket_ids, self.request.user)
        return ctx


def _order_tickets_priority_then_oldest(queryset):
    """Tri : priorité haute d'abord, puis date (plus anciens en premier)."""
    return queryset.annotate(
        _priority_order=Case(
            When(priority="high", then=Value(0)),
            When(priority="medium", then=Value(1)),
            When(priority="low", then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by("_priority_order", "created_at")


def ticket_export_excel(request):
    if not request.user.is_authenticated:
        return redirect("tickets:login")
    """Export des tickets visibles en Excel."""
    qs = get_visible_tickets_queryset(request.user)
    qs = qs.select_related("client", "member", "assigned_to")
    qs = _order_tickets_priority_then_oldest(qs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tickets"

    headers = [
        "Titre", "Client", "Membre", "Priorité", "Statut", "Type",
        "Affecté à", "Créé le", "Temps prévu (jours)", "Temps effectif (jours)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    status_labels = dict(STATUS_CHOICES)
    for row, ticket in enumerate(qs, 2):
        ws.cell(row=row, column=1, value=ticket.title)
        ws.cell(row=row, column=2, value=ticket.client.name if ticket.client else "")
        ws.cell(row=row, column=3, value=str(ticket.member) if ticket.member else "")
        ws.cell(row=row, column=4, value=ticket.get_priority_display())
        ws.cell(row=row, column=5, value=status_labels.get(ticket.status, ticket.status))
        ws.cell(row=row, column=6, value=ticket.get_type_display())
        ws.cell(row=row, column=7, value=ticket.assigned_to.display_name if ticket.assigned_to else "")
        ws.cell(row=row, column=8, value=ticket.created_at.strftime("%d/%m/%Y %H:%M") if ticket.created_at else "")
        ws.cell(row=row, column=9, value=float(ticket.estimated_time) if ticket.estimated_time else "")
        ws.cell(row=row, column=10, value=float(ticket.actual_time) if ticket.actual_time else "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tickets.xlsx"'
    wb.save(response)
    return response


class TicketDetailView(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = "tickets/ticket_detail.html"
    context_object_name = "ticket"

    def get_queryset(self):
        return get_all_tickets_queryset(self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["newly_read_comment_ids"] = self._mark_ticket_comments_as_read()
        ticket = self.object
        ctx["sent_comment_ids"] = set(
            ticket.comments.filter(author_id=self.request.user.id).values_list("id", flat=True)
        )
        ctx["comment_form"] = TicketCommentForm()
        ctx["priority_colors"] = {"low": "#198754", "medium": "#fd7e14", "high": "#dc3545"}
        ctx["can_edit"] = True
        ctx["description_attachments"] = ticket.attachments.filter(comment__isnull=True)
        return ctx

    def _mark_ticket_comments_as_read(self):
        """Marque comme lus les commentaires de l'autre camp ; retourne les IDs des commentaires nouvellement marqués (pour effet visuel)."""
        from .models import CommentReadReceipt
        from .utils import get_user_client, get_user_collaborateur, get_clients_for_collaborateur
        user = self.request.user
        ticket = self.object
        newly_read = []
        if not ticket.client:
            return newly_read
        client = get_user_client(user)
        collab = get_user_collaborateur(user)
        to_mark = []
        for c in ticket.comments.all():
            if not c.author_id:
                continue
            if client and ticket.client_id == client.id:
                if c.author_id != user.id:
                    from .models import Collaborateur
                    if Collaborateur.objects.filter(user_id=c.author_id).exists():
                        to_mark.append(c.id)
            elif collab and ticket.client_id in set(get_clients_for_collaborateur(collab).values_list("id", flat=True)):
                if ticket.client.user_id and c.author_id == ticket.client.user_id:
                    to_mark.append(c.id)
        for comment_id in to_mark:
            _, created = CommentReadReceipt.objects.get_or_create(comment_id=comment_id, user=user)
            if created:
                newly_read.append(comment_id)
        return newly_read


class TicketCreateView(LoginRequiredMixin, CreateView):
    model = Ticket
    form_class = TicketCreateForm
    template_name = "tickets/ticket_form.html"

    def dispatch(self, request, *args, **kwargs):
        if not can_create_ticket(request.user):
            return redirect("tickets:home")
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["client"] = get_user_client(self.request.user)
        kwargs["collaborateur"] = get_user_collaborateur(self.request.user)
        return kwargs

    def form_valid(self, form):
        client = get_user_client(self.request.user)
        if client:
            form.instance.client = client
        else:
            form.instance.client = form.cleaned_data.get("client")
            if not form.instance.member and form.instance.client:
                members = form.instance.client.members.all()
                if members.count() == 1:
                    form.instance.member = members.first()
        form.instance.created_by = self.request.user
        form.instance.status = "created"
        response = super().form_valid(form)
        for f in self.request.FILES.getlist("attachments"):
            TicketAttachment.objects.create(
                ticket=self.object, comment=None, file=f, name=f.name or "piece_jointe"
            )
        return response

    def get_success_url(self):
        return reverse_lazy("tickets:detail", kwargs={"pk": self.object.pk})


class TicketUpdateView(LoginRequiredMixin, UpdateView):
    model = Ticket
    form_class = TicketForm
    template_name = "tickets/ticket_form.html"
    context_object_name = "ticket"

    def get_queryset(self):
        return get_visible_tickets_queryset(self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        kwargs["client"] = get_user_client(self.request.user)
        kwargs["collaborateur"] = get_user_collaborateur(self.request.user)
        return kwargs

    def form_valid(self, form):
        if form.cleaned_data.get("status") == "validated":
            form.instance.validated_at = timezone.now()
        response = super().form_valid(form)
        for f in self.request.FILES.getlist("attachments"):
            TicketAttachment.objects.create(
                ticket=self.object, comment=None, file=f, name=f.name or "piece_jointe"
            )
        return response

    def get_success_url(self):
        return reverse_lazy("tickets:detail", kwargs={"pk": self.object.pk})


def ticket_quick_update(request, pk):
    if not request.user.is_authenticated:
        return redirect("tickets:login")
    """POST: update one field (status, priority, type, assigned_to, member, archived)."""
    ticket = get_object_or_404(Ticket, pk=pk)
    qs = get_all_tickets_queryset(request.user)
    if not qs.filter(pk=pk).exists():
        return redirect("tickets:home")
    field = request.POST.get("field")
    value = request.POST.get("value")
    if field in ("status", "priority", "type") and value:
        if field == "status" and value in dict(STATUS_CHOICES):
            ticket.status = value
            if value == "validated":
                ticket.validated_at = timezone.now()
        elif field == "priority" and value in ("low", "medium", "high"):
            ticket.priority = value
        elif field == "type" and value in ("bug", "evol", "exploit"):
            ticket.type = value
        ticket.save()
    elif field == "assigned_to" and get_user_collaborateur(request.user):
        from .models import Collaborateur
        if value == "" or value is None:
            ticket.assigned_to = None
            ticket.save()
        else:
            try:
                collab = Collaborateur.objects.filter(
                    prestataire=get_user_collaborateur(request.user).prestataire,
                    pk=value,
                ).first()
                if collab:
                    ticket.assigned_to = collab
                    ticket.save()
            except (ValueError, TypeError):
                pass
    elif field == "member" and ticket.client:
        from .models import ClientMember
        try:
            member = ClientMember.objects.filter(client=ticket.client, pk=value).first()
            if member:
                ticket.member = member
                ticket.save()
        except (ValueError, TypeError):
            pass
    elif field == "archived":
        if str(value).lower() in ("true", "1", "yes"):
            ticket.archived = True
            ticket.save()
        elif str(value).lower() in ("false", "0", "no"):
            ticket.archived = False
            ticket.save()
    return redirect(request.META.get("HTTP_REFERER", "tickets:home"))


def ticket_add_comment(request, pk):
    if not request.user.is_authenticated:
        return redirect("tickets:login")
    ticket = get_object_or_404(Ticket, pk=pk)
    if not get_visible_tickets_queryset(request.user).filter(pk=pk).exists():
        return redirect("tickets:home")
    form = TicketCommentForm(request.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        comment.ticket = ticket
        comment.author = request.user
        comment.save()
        for f in request.FILES.getlist("attachments"):
            TicketAttachment.objects.create(
                ticket=ticket, comment=comment, file=f, name=f.name
            )
        messages.success(request, "Commentaire ajouté.")
    else:
        messages.error(request, "Erreur dans le formulaire.")
    return redirect("tickets:detail", pk=pk)


def api_client_members(request, client_id):
    """Retourne les membres d'un client en JSON (pour le formulaire de création de ticket)."""
    if not request.user.is_authenticated:
        return JsonResponse({"members": []})
    collab = get_user_collaborateur(request.user)
    if not collab:
        return JsonResponse({"members": []})
    allowed_clients = get_clients_for_collaborateur(collab)
    if not allowed_clients.filter(pk=client_id).exists():
        return JsonResponse({"members": []})
    members = list(
        ClientMember.objects.filter(client_id=client_id)
        .order_by("last_name", "first_name")
        .values("id", "first_name", "last_name")
    )
    return JsonResponse({"members": members})


class StatsView(LoginRequiredMixin, View):
    """Tableau de bord statistiques (widgets sélectionnables)."""
    template_name = "tickets/stats.html"

    def get(self, request):
        qs = get_visible_tickets_queryset(request.user)
        # On inclut les archivés pour les stats "fermés"
        client = get_user_client(request.user)
        collab = get_user_collaborateur(request.user)
        if client:
            qs_all = Ticket.objects.filter(client=client)
        elif collab:
            client_ids = get_clients_for_collaborateur(collab).values_list("id", flat=True)
            qs_all = Ticket.objects.filter(client_id__in=client_ids)
        else:
            qs_all = Ticket.objects.none()

        stats_choices = request.GET.getlist("stats")
        if not stats_choices:
            stats_choices = ["open", "closed", "by_member"]

        stats = {}
        if "open" in stats_choices:
            stats["open"] = qs_all.exclude(
                status__in=["validated", "cancelled"]
            ).count()
        if "closed" in stats_choices:
            stats["closed"] = qs_all.filter(
                status__in=["validated", "cancelled"]
            ).count()
        if "by_status" in stats_choices:
            stats["by_status"] = list(
                qs_all.values("status").annotate(count=Count("id")).order_by("-count")
            )
        if "by_priority" in stats_choices:
            stats["by_priority"] = list(
                qs_all.filter(archived=False).values("priority").annotate(
                    count=Count("id")
                ).order_by("priority")
            )
        if "by_assigned" in stats_choices and collab:
            stats["by_assigned"] = list(
                qs_all.filter(archived=False)
                .filter(assigned_to__isnull=False)
                .values("assigned_to__first_name", "assigned_to__last_name")
                .annotate(count=Count("id"))
            )
        if "by_type" in stats_choices:
            stats["by_type"] = list(
                qs_all.filter(archived=False).values("type").annotate(
                    count=Count("id")
                ).order_by("type")
            )
        # Par membre (pour les clients) : ouvert / fermé par membre (validé ou annulé = fermé)
        if "by_member" in stats_choices and client:
            members = list(
                client.members.order_by("last_name", "first_name")
            )
            by_member = []
            closed_statuses = ["validated", "cancelled"]
            for m in members:
                open_count = qs_all.filter(member=m).exclude(
                    status__in=closed_statuses
                ).count()
                closed_count = qs_all.filter(member=m, status__in=closed_statuses).count()
                by_member.append({
                    "member": m,
                    "open_count": open_count,
                    "closed_count": closed_count,
                })
            stats["by_member"] = by_member

        status_labels = dict(STATUS_CHOICES)
        priority_labels = {"low": "Basse", "medium": "Moyenne", "high": "Haute"}
        type_labels = {"bug": "Bug", "evol": "Évolution", "exploit": "Exploitation"}
        return render(request, self.template_name, {
            "stats": stats,
            "stats_choices": stats_choices,
            "status_labels": status_labels,
            "priority_labels": priority_labels,
            "type_labels": type_labels,
            "user_client": client,
        })


@csrf_exempt
@require_http_methods(["POST"])
def webhook_inbound_email(request):
    """
    Webhook pour réception d'e-mails (Mailgun, SendGrid, etc.).
    Corps POST : sender, subject, body-plain (ou body_plain), Message-Id (optionnel).
    Sécurisé par EMAIL_WEBHOOK_SECRET (paramètre token ou header X-Webhook-Token).
    """
    secret = getattr(settings, "EMAIL_WEBHOOK_SECRET", None)
    if secret:
        token = request.POST.get("token") or request.headers.get("X-Webhook-Token")
        if token != secret:
            return JsonResponse({"error": "Unauthorized"}, status=401)

    # Mailgun: sender, subject, body-plain, Message-Id, pièces jointes (attachment-1, attachment-2, ...)
    from_email = request.POST.get("sender") or request.POST.get("From") or ""
    subject = request.POST.get("subject") or request.POST.get("Subject") or ""
    body = request.POST.get("body-plain") or request.POST.get("body_plain") or request.POST.get("stripped-text") or ""
    message_id = request.POST.get("Message-Id") or request.POST.get("Message-ID") or ""

    attachments = []
    for f in list(request.FILES.values())[:20]:  # max 20 PJ (aligné email_receiver)
        try:
            attachments.append((f.name or "piece_jointe", f.read()))
        except Exception:
            pass

    from .email_receiver import create_ticket_from_email

    ticket, err = create_ticket_from_email(
        from_email.strip(),
        subject,
        body,
        message_id=message_id.strip() or None,
        attachments=attachments if attachments else None,
    )
    if ticket:
        return JsonResponse({"ok": True, "ticket_id": ticket.id})
    return JsonResponse({"ok": False, "error": err or "Unknown"}, status=400)
